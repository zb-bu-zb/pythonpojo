# -*- coding: utf-8 -*- 
# @Filename : parseExcel.py           
# @Author : zhongbin
# @Time : 2023/12/8 11:17
from openpyxl import load_workbook
from datetime import datetime

class ParseExcel(object):
    '''
    解析excel文件的封装 excelPath为文件目录
    '''

    def __init__(self,excelPath):
        #操作的文件地址
        self.excelPath=excelPath
        # 使用openpyxl的函数加载excel文件到内存
        self.wb = load_workbook(self.excelPath)

    def getAllRow(self,sheetName):
        sh = self.wb[sheetName]
        size=sh.max_row
        ValueList =[]
        for i in range(1,size+1):
            ValueList.append(self.getRowValue(sheetName,i))
        return ValueList

    def getRowValue(self, sheetName, rawNo):
        '''
        获取某一行的数据
        :param sheetName:
        :param rawNo:
        :return: 列表
        '''
        sh = self.wb[sheetName]
        rowValueList = []
        for y in range(1, sh.max_column + 1):
            value = sh.cell(rawNo, y).value
            rowValueList.append(value)
        return rowValueList

    def getColumnValue(self, sheetName, colNo):
        '''
        获取某一列的数据
        :param sheetName:
        :param colNo:
        :return: 列表
        '''
        sh = self.wb[sheetName]
        colValueList = []
        for x in range(2, sh.max_row + 1):
            value = sh.cell(x, colNo).value
            colValueList.append(value)
        return colValueList

    def getCellOfValue(self, sheetName, rowNo, colNo):
        '''
        获取某一个单元格的数据
        :param sheetName:单元格名称
        :param rowNo:行
        :param colNo:列
        :return: 字符串
        '''
        sh = self.wb[sheetName]
        value = sh.cell(rowNo, colNo).value
        return value

    def writeCell(self, sheetName, rowNo, colNo, value):
        '''
        向某个单元格写入数据
        :param rowNo: 行号
        :param colNo: 列号
        :param value:
        :return: 无
        '''
        sh = self.wb[sheetName]
        sh.cell(rowNo, colNo).value = value
        self.wb.save(self.excelPath)

    def writeCurrentTime(self, sheetName, rowNo, colNo):
        '''
        向某个单元格写入当前时间
        :return:
        '''
        sh = self.wb[sheetName]
        Time = datetime.now()
        currentTime = Time.strftime('%Y:%m:%d %H:%M:%S')
        sh.cell(rowNo, colNo).value = currentTime
        self.wb.save(self.excelPath)




